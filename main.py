from fastapi import FastAPI, File, UploadFile, HTTPException, Request
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import cv2
import numpy as np
from ultralytics import YOLO
import logging
import os
import sqlite3
from datetime import datetime
import uuid
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import tempfile

# Инициализация приложения FastAPI
app = FastAPI(
    title="Vehicle Detection API",
    description="API для подсчёта автомобилей на изображениях с использованием YOLOv8"
)

# Создание директорий, если их нет
os.makedirs("static/results", exist_ok=True)
os.makedirs("templates", exist_ok=True)
os.makedirs("database", exist_ok=True)

# Настройка путей для статических файлов и шаблонов
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Путь к базе данных
DATABASE_PATH = "database/vehicle_detection.db"

# Инициализация базы данных
def init_db():
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS detections (
        id TEXT PRIMARY KEY,
        filename TEXT NOT NULL,
        upload_time TIMESTAMP NOT NULL,
        status TEXT NOT NULL,
        error_message TEXT,
        total_vehicles INTEGER,
        cars INTEGER,
        buses INTEGER,
        trucks INTEGER,
        motorcycles INTEGER,
        result_image_path TEXT
    )
    ''')
    conn.commit()
    conn.close()
    logging.info("Database initialized")

# Вызов инициализации БД при старте
init_db()

# Загрузка модели YOLOv8
model = YOLO('yolov8n.pt')  # Модель nano (самая быстрая)

# Настройка логгера
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Классы транспортных средств в COCO Dataset
CLASS_NAMES = {
    2: "car",
    3: "motorcycle",
    5: "bus",
    7: "truck"
}
VEHICLE_CLASSES = CLASS_NAMES.keys()

# Максимальный размер парковки по умолчанию
max_parking_size = 50

def generate_excel_report(detections):
    """Генерирует Excel-файл с отчётом о всех детекциях"""
    try:
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehicle Detections"
        
        # Заголовки столбцов
        headers = [
            "ID", "Имя файла", "Время загрузки", "Статус", 
            "Всего ТС", "Автомобили", "Автобусы", "Грузовики", "Мотоциклы",
            "Ошибка", "Путь к изображению"
        ]
        
        # Стили для оформления
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='thin'))
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment
        
        # Заполняем данными
        for row_num, detection in enumerate(detections, 2):
            ws.cell(row=row_num, column=1, value=detection['id'])
            ws.cell(row=row_num, column=2, value=detection['filename'])
            ws.cell(row=row_num, column=3, value=detection['upload_time'])
            ws.cell(row=row_num, column=4, value=detection['status'])
            
            # Для успешных запросов добавляем цифры
            if detection['status'] == 'success':
                ws.cell(row=row_num, column=5, value=detection['total_vehicles'])
                ws.cell(row=row_num, column=6, value=detection['cars'])
                ws.cell(row=row_num, column=7, value=detection['buses'])
                ws.cell(row=row_num, column=8, value=detection['trucks'])
                ws.cell(row=row_num, column=9, value=detection['motorcycles'])
                ws.cell(row=row_num, column=10, value="")
            else:
                # Для ошибок оставляем пустые значения в столбцах с цифрами
                for col in range(5, 10):
                    ws.cell(row=row_num, column=col, value="")
                ws.cell(row=row_num, column=10, value=detection['error_message'])
            
            ws.cell(row=row_num, column=11, value=detection['result_image_path'])
        
        # Настраиваем ширину столбцов
        column_widths = [36, 30, 20, 10, 10, 12, 12, 12, 12, 50, 50]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Применяем стили ко всем ячейкам

        for row in ws.iter_rows(min_row=1, max_row=len(detections)+1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                cell.alignment = alignment
        
        # Создаем временный файл
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    except Exception as e:
        logger.error(f"Ошибка генерации Excel: {str(e)}")
        return None

@app.get("/export-excel/", summary="Экспорт всех детекций в Excel")
async def export_to_excel():
    """Возвращает Excel-файл со всеми записями о детекциях"""
    try:
        detections = get_all_detections()
        if not detections:
            raise HTTPException(status_code=404, detail="Нет данных для экспорта")
        
        excel_path = generate_excel_report(detections)
        if not excel_path:
            raise HTTPException(status_code=500, detail="Ошибка генерации отчёта")
        
        # Имя файла с текущей датой
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"vehicle_detections_{date_str}.xlsx"
        
        return FileResponse(
            excel_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        logger.error(f"Ошибка экспорта в Excel: {str(e)}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

def save_detection_to_db(detection_data: dict):
    """Сохраняет результаты обнаружения в базу данных"""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        cursor.execute('''
        INSERT INTO detections (
            id, filename, upload_time, status, error_message,
            total_vehicles, cars, buses, trucks, motorcycles, result_image_path
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            detection_data['id'],
            detection_data['filename'],
            detection_data['upload_time'],
            detection_data['status'],
            detection_data.get('error_message'),
            detection_data.get('total_vehicles'),
            detection_data.get('cars'),
            detection_data.get('buses'),
            detection_data.get('trucks'),
            detection_data.get('motorcycles'),
            detection_data.get('result_image_path')
        ))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logger.error(f"Ошибка сохранения в БД: {str(e)}")
        return False

def get_all_detections():
    """Получает все записи из базы данных"""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM detections ORDER BY upload_time DESC")
        detections = cursor.fetchall()
        conn.close()
        
        # Преобразуем в список словарей
        columns = ['id', 'filename', 'upload_time', 'status', 'error_message',
                   'total_vehicles', 'cars', 'buses', 'trucks', 'motorcycles', 'result_image_path']
        return [dict(zip(columns, row)) for row in detections]
    except Exception as e:
        logger.error(f"Ошибка чтения из БД: {str(e)}")
        return []

def save_error_to_db(detection_id, filename, upload_time, error_message):
    """Сохраняет информацию об ошибке в базу данных"""
    detection_data = {
        'id': detection_id,
        'filename': filename,
        'upload_time': upload_time.isoformat(),
        'status': 'error',
        'error_message': error_message,
        'total_vehicles': None,
        'cars': None,
        'buses': None,
        'trucks': None,
        'motorcycles': None,
        'result_image_path': None
    }
    save_detection_to_db(detection_data)

@app.get("/", response_class=HTMLResponse, include_in_schema=False)
async def read_root(request: Request):
    """
    Возвращает HTML-страницу для загрузки изображения
    """
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/history", response_class=HTMLResponse, include_in_schema=False)
async def detection_history(request: Request):
    """Страница с историей обнаружений"""
    detections = get_all_detections()
    return templates.TemplateResponse("history.html", {
        "request": request,
        "detections": detections
    })

@app.post("/detect-vehicles/", summary="Обнаружение транспортных средств")
async def detect_vehicles(
    request: Request,
    file: UploadFile = File(..., description="Изображение для анализа")
):
    """
    Принимает изображение и возвращает количество обнаруженных транспортных средств.
    Поддерживает как API-запросы (JSON), так и веб-формы (HTML).
    """
    # Генерация уникального ID для этого запроса
    detection_id = str(uuid.uuid4())
    upload_time = datetime.utcnow()
    filename = file.filename if file.filename else "unknown"
    
    try:
        # Проверка формата файла
        if file.content_type not in ["image/jpeg", "image/jpg", "image/png"]:
            error_msg = "Неподдерживаемый формат изображения. Используйте JPEG или PNG."
            save_error_to_db(detection_id, filename, upload_time, error_msg)
            
            if "text/html" in request.headers.get("accept", ""):
                return templates.TemplateResponse("index.html", {
                    "request": request,
                    "error": error_msg
                })
            raise HTTPException(status_code=400, detail=error_msg)

        # Чтение изображения
        contents = await file.read()
        nparr = np.frombuffer(contents, np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if image is None:
            error_msg = "Ошибка декодирования изображения"
            save_error_to_db(detection_id, filename, upload_time, error_msg)
            
            if "text/html" in request.headers.get("accept", ""):
                return templates.TemplateResponse("index.html", {
                    "request": request,
                    "error": error_msg
                })
            raise HTTPException(status_code=400, detail=error_msg)

        # Детекция объектов
        results = model(image, verbose=False)
        
        # Инициализация счётчиков
        vehicle_counts = {'car': 0, 'bus': 0, 'truck': 0, 'motorcycle': 0}
        total_vehicles = 0
        detection_image = image.copy()

        # Обработка результатов
        for result in results:
            for box in result.boxes:
                class_id = int(box.cls)
                if class_id in VEHICLE_CLASSES:
                    vehicle_class = CLASS_NAMES[class_id]
                    vehicle_counts[vehicle_class] += 1
                    total_vehicles += 1
                    
                    # Рисуем bounding box
                    x1, y1, x2, y2 = map(int, box.xyxy[0])
                    cv2.rectangle(detection_image, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.putText(detection_image, vehicle_class, (x1, y1 - 10), 
                                cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

        # Сохраняем изображение с детекциями
        result_filename = f"results/{detection_id}_{filename}"
        result_path = f"static/{result_filename}"
        os.makedirs(os.path.dirname(result_path), exist_ok=True)
        cv2.imwrite(result_path, detection_image)

        # Подготавливаем данные для сохранения в БД
        detection_data = {
            'id': detection_id,
            'filename': filename,
            'upload_time': upload_time.isoformat(),
            'status': 'success',
            'error_message': None,
            'total_vehicles': total_vehicles,
            'cars': vehicle_counts['car'],
            'buses': vehicle_counts['bus'],
            'trucks': vehicle_counts['truck'],
            'motorcycles': vehicle_counts['motorcycle'],
            'result_image_path': f"/static/{result_filename}"
        }
        
        # Сохраняем в базу данных
        save_detection_to_db(detection_data)

        # Формирование ответа
        response_data = {
            "id": detection_id,
            "total_vehicles": total_vehicles,
            "details": vehicle_counts,
            "message": f"Найдено {total_vehicles} транспортных средств",
            "image_url": f"/static/{result_filename}",
            "filename": filename
        }
        
        # Возвращаем HTML для веб-интерфейса или JSON для API
        if "text/html" in request.headers.get("accept", ""):
            return templates.TemplateResponse("result.html", {
                "request": request,
                "max_parking_size": max_parking_size,
                **response_data,
            })
        
        return JSONResponse(content=response_data)
    
    except Exception as e:
        # Получаем полный стектрейс ошибки
        error_trace = traceback.format_exc()
        error_msg = f"Внутренняя ошибка сервера: {str(e)}"
        logger.error(f"Ошибка обработки изображения: {error_trace}")
        
        # Сохраняем информацию об ошибке в БД
        save_error_to_db(detection_id, filename, upload_time, error_msg)
        
        if "text/html" in request.headers.get("accept", ""):
            return templates.TemplateResponse("index.html", {
                "request": request,
                "error": error_msg
            })
        raise HTTPException(status_code=500, detail=error_msg)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
